from MySqlConector import connectToHost, closeConectionToHost, executeQuery
from letrasanumeros import numaLetras
import pandas as pd
import openpyxl
from datetime import date

def createExcelInforme(hostsql, usersql, passwordsql, dataBaseName, tableName, key, codigoTasa, cliente, fechaInspecc, tcTasa, formatFile, mainDataFilePath, formatReportfile):
    
    connection, cursor = connectToHost(hostsql, usersql, passwordsql) # Init connection

    # Use dataBase
    data = None
    menssage = f"You're connected to database: {dataBaseName}!"
    command = f'USE {dataBaseName};'
    query = (command, data, menssage)
    executeQuery(connection, cursor, dataBaseName, query)

    DataOutMatrix = pd.read_sql(F'SELECT * FROM {tableName}', connection)

    closeConectionToHost(connection, cursor) # Close connection

    # Creates dataframe with the data from selected real estate units:
    tasacionNew = pd.DataFrame(columns= DataOutMatrix.columns)
    for x in key:
        tasacionNew = tasacionNew.append(DataOutMatrix.loc[key[x]])
    
    # Organize data so it can be analyze and upload to the report excel form
    unidad = tasacionNew.iloc[:]['unidad']
    num = tasacionNew.iloc[:]['num']
    nivel = tasacionNew.iloc[:]['nivel']
    aTerreno = tasacionNew['terreno']
    aTechada = tasacionNew['techada']
    aOcupada = tasacionNew['ocupada']
    descrip = tasacionNew['descripción']
    vComercialUsd = tasacionNew['comercialusd']
    vRealizaUsd = tasacionNew['realizausd']

    # Get bank name and comparse to get letters:
    wbProyecto = openpyxl.load_workbook(formatFile, read_only=True)
    sheetInforme = wbProyecto['Memoria']
    banco = sheetInforme['D7'].value
    sheetPortada = wbProyecto['Portada']
    projectName = sheetPortada['B33'].value
    wbProyecto.close
    if banco == 'BANCO INTERNACIONAL DEL PERÚ S.A.A. - INTERBANK':
        banco = 'IBK'
    elif banco == 'BANCO PINCHINCHA':
        banco = 'BP'
    elif banco == 'SCOTIABANK PERÚ S.A.A.':
        banco = 'SCB'

    codigoTasa = f'Informe de Valuación N° {codigoTasa}-{banco}-{str(date.today().strftime("%Y"))}' # Cretaes report number code

    # Creates tipo string:
    tipo = ''
    for x in range(0, len(key), 1):
        if x == 0:
            tipo = unidad[key[x]]
        else:
            if not unidad[key[x]] == unidad[key[x-1]]:
                tipo = f"{tipo}, {unidad[key[x]]}"

    # Gets VUE and VRT values:
    DataInGeneral = pd.read_excel(mainDataFilePath, sheet_name='DATA')
    vut = float(DataInGeneral.iloc[:1]['VUT (USD)']) #Unit terrain value
    vrc = float(DataInGeneral.iloc[:1]['VRC (USD)']) #Unit reconstruction value

    # Opens main report file:
    wbInforme = openpyxl.load_workbook(formatReportfile, read_only=False)
    sheetInforme = wbInforme['Informe']

    # Fills main report data:
    sheetInforme['B4'].value = codigoTasa
    sheetInforme['I7'].value = cliente
    sheetInforme['I12'].value = fechaInspecc
    sheetInforme['E67'].value = tcTasa
    sheetInforme['J74'].value = tipo

    
    # Fills real estate units data in report:
    y = 0
    vComercialUsdTotal = 0
    vRealizaUsdTotal = 0
    for x in range(0, len(key), 1):
        sheetInforme.cell(row=x + 138, column=4).value = f"{unidad[key[x]]} No {num[key[x]]}"
        sheetInforme.cell(row=x + 138, column=16).value = aTerreno[key[x]]
        sheetInforme.cell(row=x + 154, column=11).value = aTechada[key[x]]
        sheetInforme.cell(row=x + 154, column=16).value = aOcupada[key[x]]
        sheetInforme.cell(row=x + 189, column=4).value = f"Es materia de tasación el {unidad[key[x]]} No {num[key[x]]}, ubicado en el nivel {nivel[key[x]]}"
        sheetInforme.cell(row=x + 200, column=13).value = descrip[key[x]]
        sheetInforme.cell(row=x + 287, column=14).value = vut
        sheetInforme.cell(row=x + 324, column=17).value = vrc
        sheetInforme.cell(row=x + 340, column=19).value = vComercialUsd[key[x]]
        sheetInforme.cell(row=x + 366, column=15).value = round(vRealizaUsd[key[x]] / vComercialUsd[key[x]], 1)
        vComercialUsdTotal = vComercialUsdTotal + vComercialUsd[key[x]]
        vRealizaUsdTotal = vRealizaUsdTotal + vRealizaUsd[key[x]]
        y = y + 1

    # Hides excess rows for printing:
    filasInforme = [376, 350, 334, 316, 297, 210, 199, 180, 164, 148, 110, 65, 51]
    for x in filasInforme:
        for n in range(0, 11-y, 1):
            sheetInforme.row_dimensions[x-n].hidden = True

    letras = numaLetras(vComercialUsdTotal)
    letras1 = numaLetras(vRealizaUsdTotal)
    sheetInforme['F353'].value = f'{letras} DÓLARES AMERICANOS'
    sheetInforme['F379'].value = f'{letras1} DÓLARES AMERICANOS'

    #Saves and closes the new report:
    wbInforme.save(f'{str(date.today().strftime("%Y-%m-%d"))}_{projectName}_{cliente}.xlsx')
    wbInforme.close


if __name__ == '__main__':

    # Server connection informartion:   
    hostsql = 'localhost'
    usersql = 'root'
    passwordsql = 'acidbass'

    dataBaseName = 'proyecto1'#input('Nombre de base de datos: ')
    tableName = 'matriz' # Table with project matrix data

    formatFile = 'Formatos\PMF-Tasacion.xlsx'
    mainDataFilePath = 'Formatos\DATA.xlsx'
    formatReportfile = 'Formatos\Informe.xlsx'

    key = {0:1, 1:183, 2:224} # User input selecction of real estate units

    # User input:
    codigoTasa = '256' 
    cliente = 'Pepito Mendienta - DNI 45768939'
    fechaInspecc = '20/10/2020'
    tcTasa = float('3.6')

    createExcelInforme(hostsql, usersql, passwordsql, dataBaseName, tableName, key, codigoTasa, cliente, fechaInspecc, tcTasa, formatFile, mainDataFilePath, formatReportfile)