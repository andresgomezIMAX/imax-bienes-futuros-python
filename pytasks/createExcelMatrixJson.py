from jsonConverter import recibeJson
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Color, colors, PatternFill
from datetime import date

def createExcelMatrix(formatMatrixFile, formatFile, pathmatrix, pathExcel):

    # Converts Json to DataFrame
    DataOutMatrix = recibeJson(pathmatrix)
    DataOutMatrix = DataOutMatrix.fillna('')

    #Organize data so it can be analyze and upload to the matrix excel form
    unidad = DataOutMatrix.iloc[:]['unidad']
    num = DataOutMatrix.iloc[:]['num']
    nivel = DataOutMatrix.iloc[:]['nivel']
    dni = DataOutMatrix.iloc[:]['dni']
    nombre = DataOutMatrix.iloc[:]['nombre']
    aTerreno = DataOutMatrix.iloc[:]['terreno']
    aOcupada = DataOutMatrix.iloc[:]['ocupada']
    aTechada = DataOutMatrix.iloc[:]['techada']
    aComunes = DataOutMatrix.iloc[:]['comunes']
    moneda = DataOutMatrix.iloc[:]['moneda']
    vVenta = DataOutMatrix.iloc[:]['valor']
    vTerrenoUsd = DataOutMatrix.iloc[:]['terrenousd']
    vEdificaUsd = DataOutMatrix.iloc[:]['edifusd']
    vComercialUsd = DataOutMatrix.iloc[:]['comercialusd']
    vRealizaUsd = DataOutMatrix.iloc[:]['realizausd']
    vAseguraUsd = DataOutMatrix.iloc[:]['asegurausd']
    tc = DataOutMatrix.iloc[:]['tipocambio']
    fecha = DataOutMatrix.iloc[:]['fecha']

    wbProyecto = openpyxl.load_workbook(formatFile, read_only=True) #Opens form report file read only
    sheetProyecto = wbProyecto['Portada']

    wbMatrix = openpyxl.load_workbook(formatMatrixFile, read_only=False) #Opens Matrix form file so it can be updated
    sheetMatrix = wbMatrix['MATRIZ']

    for x in range(0, len(DataOutMatrix), 1):

        # Calculates data in Peruvian currency - Round numbers
        vTerrenoSol = round(float(vTerrenoUsd[x]) * float(tc[x]) / 100, 0) * 100
        vEdificaSol = round(float(vEdificaUsd[x]) * float(tc[x]) / 100, 0) * 100
        vComercialSol = round(float(vComercialUsd[x]) * float(tc[x]) / 100, 0) * 100
        vRealizaSol = round(float(vRealizaUsd[x]) * float(tc[x]) / 100, 0) * 100
        vAseguraSol = round(float(vAseguraUsd[x]) * float(tc[x]) / 100, 0) * 100

        rowData = [unidad[x], num[x], nivel[x], dni[x], nombre[x], aTerreno[x], aOcupada[x],
                    aTechada[x], aComunes[x], moneda[x], vVenta[x], aTerreno[x], aOcupada[x], 
                    aTechada[x], aComunes[x], vTerrenoUsd[x], vEdificaUsd[x], vComercialUsd[x],
                    vRealizaUsd[x], vAseguraUsd[x], tc[x], vTerrenoSol, vEdificaSol, vComercialSol,
                    vRealizaSol, vAseguraSol, fecha[x]]

        sheetMatrix.insert_rows(x+13) # Inserts rows

        #Insert data in rows:
        for y in range(0, len(rowData), 1):
            sheetMatrix.cell(row=x+12, column=y+1).value = rowData[y]
            
            # Correction of the rows style:
            if sheetMatrix.cell(row=x+12, column=y+1).has_style:
                sheetMatrix.cell(row=x+13, column=y+1)._style = sheetMatrix.cell(row=x+12, column=y+1)._style
            
            if not fecha[x] == '':
                sheetMatrix.cell(row=x+12, column=y+1).font = Font(bold=True)
                sheetMatrix.cell(row=x+12, column=y+1).fill = PatternFill("solid", fgColor="DDEBF7") 

    # Updates formulas in excel
    sheetMatrix[f'F{x+14}'] = f'=SUM(F12:F{x+13})'
    sheetMatrix[f'G{x+14}'] = f'=SUM(G12:G{x+13})'
    sheetMatrix[f'H{x+14}'] = f'=SUM(H12:H{x+13})'
    sheetMatrix[f'I{x+14}'] = f'=SUM(I12:I{x+13})'
    sheetMatrix['D7'] = f'=MAX(AA12:AA{x+13})'
    sheetMatrix['D8'] = str(date.today().strftime("%d-%m-%Y"))

    # Uodates general project Data:
    developerName = sheetProyecto['B34'].value
    sheetMatrix['D3'].value = developerName

    projectName = sheetProyecto['B33'].value
    sheetMatrix['D5'].value = projectName

    sheetProyecto = wbProyecto['Memoria']
    sheetMatrix['D6'].value = sheetProyecto['D31'].value

    wbMatrix.save(f'{pathExcel}{str(date.today().strftime("%Y-%m-%d"))}_Matriz_{projectName}.xlsx') # Saves matrix excel

    wbProyecto.close
    wbMatrix.close

if __name__ == '__main__':

    projects = recibeJson('Formatos\Template.json')
    projectSelected = 0

    formatFile = projects['PathInforme'][projectSelected] # Excel with project data
    
    formatMatrixFile = 'Formatos\MATRIZ TASACIONES.xlsx' # Excel with Matrix Form
    
    pathmatrix = 'matrixjson.json'

    pathExcel = ''

    createExcelMatrix(formatMatrixFile, formatFile, pathmatrix, pathExcel)