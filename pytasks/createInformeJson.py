from letrasanumeros import numaLetras
from createMatrixJson import calculoTasa
from jsonConverter import recibeJson, sentJson
import pandas as pd
import openpyxl
from datetime import date

def createExcelInforme(pathdatauitasa, formatFile, mainDataFilePath, formatReportfile, pathmatrix, pathdatageneral, pathExcel):

    fecha = str(date.today().strftime("%Y-%m-%d"))

    # 'fila', 'valor', 'aOcupada', 'aTechada', 'factorCalculo'
    dataUITasa = recibeJson(pathdatauitasa)
    key = dataUITasa['fila']
    dataUITasa = dataUITasa.rename(index = dataUITasa.iloc[:]['fila'])
    aTechada = dataUITasa['aTechada']
    aOcupada = dataUITasa['aOcupada']
    factorTasacion = dataUITasa['factorCalculo']
    moneda = dataUITasa['moneda']
    vVenta = dataUITasa['valor']

    # 'código', 'nombre', 'dni', 'conyugue', 'dniC', 'fechaInspec', 'perito', 'tipoCambio'
    dataGeneralTasa = recibeJson(pathdatageneral)
    codigoTasa = dataGeneralTasa['código'][0]
    clienteFile = f"{dataGeneralTasa['nombre'][0]} - DNI {dataGeneralTasa['dni'][0]}"
    cliente = f"{dataGeneralTasa['nombre'][0]} - DNI: {dataGeneralTasa['dni'][0]} / {dataGeneralTasa['conyugue'][0]} - DNI: {dataGeneralTasa['dniC'][0]}"
    fechaInspecc = dataGeneralTasa['fechaInspec'][0]
    tcTasa = dataGeneralTasa['tipoCambio'][0]
    perito = dataGeneralTasa['perito'][0]
    dni = f"{dataGeneralTasa['dni'][0]} / {dataGeneralTasa['dniC'][0]}"
    nombre = f"{dataGeneralTasa['nombre'][0]} / {dataGeneralTasa['conyugue'][0]}"

    # Converts Json to DataFrame
    DataOutMatrix = recibeJson(pathmatrix)
    DataOutMatrix = DataOutMatrix.fillna('')
    totalesData = DataOutMatrix.sum(axis=0, skipna=True) #Sum all data in columns
    aOcupadaSum = totalesData['ocupada']

    # Creates dataframe with the data from selected real estate units:
    tasacionNew = pd.DataFrame(columns= DataOutMatrix.columns)
    for x in key:
        tasacionNew = tasacionNew.append(DataOutMatrix.loc[x])

    # Organize data so it can be analyze and upload to the report excel form
    unidad = tasacionNew.iloc[:]['unidad']
    num = tasacionNew.iloc[:]['num']
    nivel = tasacionNew.iloc[:]['nivel']
    clase = tasacionNew.iloc[:]['clase']
    claseUI = tasacionNew[['tipo', 'descripción']]
    claseUI = claseUI.rename(index = tasacionNew.iloc[:]['clase']) #Gets type of real estate unit data
    
    # Get bank name and comparse to get letters:
    wbProyecto = openpyxl.load_workbook(formatFile, read_only=True)
    sheetInforme = wbProyecto['Memoria']
    eFinan = sheetInforme['D7'].value
    sheetPortada = wbProyecto['Portada']
    projectName = sheetPortada['B33'].value
    aTerrenoPredio = float(sheetInforme['P94'].value)
    wbProyecto.close
    
    if eFinan == 'BANCO INTERNACIONAL DEL PERÚ S.A.A. - INTERBANK':
        banco = 'IBK'
    elif eFinan == 'BANCO PINCHINCHA':
        banco = 'BP'
    elif eFinan == 'SCOTIABANK PERÚ S.A.A.':
        banco = 'SCB'

    codigoTasa = f'Informe de Valuación N° {codigoTasa}-{banco}-{str(date.today().strftime("%Y"))}' # Cretaes report number code

    # Creates tipo string:
    tipo = ''
    for x in range(0, len(key), 1):
        if x == 0:
            tipo = unidad[key[x]]
        else:
            if not unidad[key[x]] == unidad[key[x-1]]:
                tipo = f"{tipo}, {unidad[key[x]]}"

    # Gets VUE and VRT values:
    DataInGeneral = pd.read_excel(mainDataFilePath, sheet_name='DATA')
    vut = float(DataInGeneral.iloc[:1]['VUT (USD)']) #Unit terrain value
    vrc = float(DataInGeneral.iloc[:1]['VRC (USD)']) #Unit reconstruction value
    factorRealiza = float(DataInGeneral.iloc[:1]['F.Realiza (%)']) #Realization value factor for calculus
    factorAsegura = float(DataInGeneral.iloc[:1]['F.Asegura (%)']) #Insurable value factor for calculus
    aComunTotal = float(DataInGeneral.iloc[:1]['A. Común (m²)']) #Projects Total Comun area
    aOcupadaTotal = aOcupadaSum + aComunTotal #Total area including comun area

    # Opens main report file:
    wbInforme = openpyxl.load_workbook(formatReportfile, read_only=False)
    sheetInforme = wbInforme['Informe']

    # Fills main report data:
    sheetInforme['B4'].value = codigoTasa
    sheetInforme['I7'].value = cliente
    sheetInforme['I12'].value = fechaInspecc
    sheetInforme['E67'].value = tcTasa
    sheetInforme['J74'].value = tipo

    #Creates tasaciones Table:
    columnas = [
                'unidad', 'num', 'nivel', 'dni', 'nombre',
                'terreno', 'ocupada', 'techada', 'comunes', 'moneda', 'valor', 
                'terrenousd', 'edifusd', 'comercialusd', 'realizausd',
                'asegurausd', 'tipocambio', 'fecha', 'clase', 'vueusd', 'tipo', 'descripción'
                ]

    tasacionSegui = pd.DataFrame(columns= columnas)


    # Fills real estate units data in report:
    y = 0
    vComercialUsdTotal = 0
    vRealizaUsdTotal = 0
    for x in dataUITasa.index:

        factorT = factorTasacion[x]
        
        data = calculoTasa(aComunTotal, aOcupadaTotal, aTerrenoPredio, aOcupada[x], aOcupadaSum, vut, factorT, unidad[x], moneda[x], vVenta[x], aTechada[x], factorRealiza, eFinan, vrc, factorAsegura, tcTasa, claseUI, num[x], nivel[x], clase[x], dni, nombre, fecha)

        sheetInforme.cell(row=y + 138, column=4).value = f"{unidad[x]} No {num[x]}"
        sheetInforme.cell(row=y + 138, column=16).value = data['terreno']
        sheetInforme.cell(row=y + 154, column=11).value = aTechada[x]
        sheetInforme.cell(row=y + 154, column=16).value = aOcupada[x]
        sheetInforme.cell(row=y + 189, column=4).value = f"Es materia de tasación el {unidad[x]} No {num[x]}, ubicado en el nivel {nivel[x]}"
        sheetInforme.cell(row=y + 200, column=13).value = data['descripción']
        sheetInforme.cell(row=y + 287, column=14).value = vut
        sheetInforme.cell(row=y + 324, column=17).value = vrc
        sheetInforme.cell(row=y + 340, column=19).value = data['comercialusd']
        sheetInforme.cell(row=y + 366, column=15).value = round(data['realizausd'] / data['comercialusd'], 1)
        vComercialUsdTotal = vComercialUsdTotal + data['comercialusd']
        vRealizaUsdTotal = vRealizaUsdTotal + data['realizausd']

        tasacionSegui = tasacionSegui.append(data, ignore_index=True)

        y = y + 1

    tasacionSegui = tasacionSegui.rename(index = key)
    DataOutMatrix.loc[key] = tasacionSegui

    # Hides excess rows for printing:
    filasInforme = [376, 350, 334, 316, 297, 210, 199, 180, 164, 148, 110, 65, 51]
    for x in filasInforme:
        for n in range(0, 11-y, 1):
            sheetInforme.row_dimensions[x-n].hidden = True

    letras = numaLetras(vComercialUsdTotal)
    letras1 = numaLetras(vRealizaUsdTotal)
    sheetInforme['F353'].value = f'{letras} DÓLARES AMERICANOS'
    sheetInforme['F379'].value = f'{letras1} DÓLARES AMERICANOS'

    #Saves and closes the new report:
    wbInforme.save(f'{pathExcel}{fecha}_{projectName}_{clienteFile}.xlsx')
    wbInforme.close

    return tasacionSegui, DataOutMatrix

if __name__ == '__main__':

    projects = recibeJson('Formatos\Template.json')
    projectSelected = 0

    mainDataFilePath = projects['PathData'][projectSelected] # Excel with matrix data
    formatFile = projects['PathInforme'][projectSelected] # Excel with project data

    formatReportfile = 'Formatos\Informe.xlsx' #Ubicación de formato para informe

    pathmatrix = 'matrixjson.json'
    pathdatageneral = 'dataGeneral.json'
    pathdatauitasa = 'dataUITasa.json'

    pathExcel = ''

    tasacionSegui, DataOutMatrix = createExcelInforme(pathdatauitasa, formatFile, mainDataFilePath, formatReportfile, pathmatrix, pathdatageneral, pathExcel)

    nameT = 'tasaciones.json'
    nameM = 'matrixUpdatejson.json'

    sentJson(tasacionSegui, nameT)
    sentJson(DataOutMatrix, nameM)
